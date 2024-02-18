
#imports
import tkinter as tk
from tkinter import filedialog
from itertools import zip_longest
import binascii
from PIL import ImageTk, Image
import operator 
import xml.etree.cElementTree as ET
import xml.dom.minidom
import openpyxl

#Global defined values
line_start_with = " "  #For skipping the header
line_divider = chr(0x255F)+chr(0x2500)*9+chr(0x253C)+chr(0x2500)*19+chr(0x253C)+chr(0x2500)*19+chr(0x2562)+ "\n"
global list_error 


#The file paths of the two .trc files to compare
global file_path 
#Parsed CAN messages
global parsed_messages
#Structure sortie comparaison
global compared_msg
#list final
global list_final

#Class TrcParser for parsing messages in the .trc file
class TrcParser:
    def __init__(self, filename):

        self.messages = []
        #count total number of messages in file
        number_lines_header = 0

        with open(filename, "r") as f:
            #skip header lines until reaching data
            while True:
                header = f.readline()
                if header.startswith(line_start_with):
                    break
                else :
                    number_lines_header += 1
        with open(filename, 'r') as file:
            lines = sum(1 for line in file)
            self.nb_messages = lines-number_lines_header #number of messages (total number of lines - number of lines in header)
        with open(filename, "r") as f:
            while True:
                line = f.readline()
                if line.startswith(line_start_with):
                    break
            idx=0
            #Read CAN messages until end of file
            while True:
                print("reading")
                if not line:
                    break

                line = line.strip() 
                elements = line.split() 
                #dlc = elements[4]
                try:
                    if elements[5] == "_" or elements[5] == "-":  
                        data_list = elements[7:]
                        data = ''.join(data_list)
                        self.messages.append({
                                "Timestamp": elements[1],
                                "ID": elements[4],
                                "Data": [byte for byte in bytes.fromhex(data)]   
                        })
                    else: 
                        data_list = elements[5:]
                        data = ''.join(data_list)
                        self.messages.append({
                                "Timestamp": elements[1],
                                "ID": elements[3],
                                "Data": [byte for byte in bytes.fromhex(data)]   
                        })
                except:
                    raise
                    print("error")
                line = f.readline()
            #print("debug")


def open_file():
    global file_path 
    global parsed_messages
    global file_1_set
    # Utilise tkinter pour ouvrir le file dialog
    root = tk.Tk()
    root.withdraw()  # Cache la fenêtre principale
    file_path = filedialog.askopenfilename()

    if file_path :
        file_1_set = True
        print("Selected file:", file_path)
        parsed_messages = TrcParser(file_path)
        pass
    #print("debug")

def get_time_data():
    global parsed_messages #True data
    global list_final
 
    #Tri les datas trouvées par odre chronologique TimeTrue car tmeps avec les us
    #compared_msg.sort(key=operator.itemgetter('TimeTrue'))
    data_per_id = [] #stockage temporaire des data par id
    list_data_per_id =[]
    list_id = []

    #Fait une liste de tout les id
    for msg in parsed_messages.messages:
        id_finded = False
        for id in list_id:
            if id == msg["ID"] :
                id_finded = True
                break
        if  id_finded == False:
            list_id.append(msg["ID"])
    
    #Tri les datas pour le regrouper par id
    for id in list_id:
        for data in parsed_messages.messages:
            if id == data["ID"] :
                data_per_id.append(data)
        list_data_per_id.append(data_per_id)
        data_per_id = []
    
    #calcul les fréquences
    list_frequences = []
    for id in list_data_per_id: #parcour la liste de data
        list_time = []
        old_time = 0
        time_offset = 0
        time_offset_max = 0
        time_offset_min = 1000000000 
        t_moy = 0
        for position, data in enumerate(id): #parcours chaque liste d'id
            #calcul temps
            if position > 0: #pas de calcul pour la première valeure
                #time_offset = round(float(data["TimeTrue"]) - old_time,2)
                time_offset = round(float(data["Timestamp"]) - old_time,2)
                list_time.append(time_offset)
            #old_time = round(float(data["TimeTrue"]),2) #mise en mémoire
            old_time = round(float(data["Timestamp"]),2)
            #calcul max
            if time_offset > time_offset_max:
                time_offset_max = time_offset
            #calcul min
            if position > 0:
                if time_offset < time_offset_min:
                    time_offset_min = time_offset
        #calcul temps moyen
        if len(list_time) > 0:
            t_moy = round(sum(list_time)/len(list_time),2)
            #créer liste fréquence
            #les id sont convertie car on a doit les avoir en decimal pour comparer avec le xml
            list_frequences.append({"id": int(data["ID"],16),"t_moy": t_moy , "t_max": time_offset_max ,"t_min": time_offset_min} )
    
    #Génération liste finale
    list_final = []
    for position, data in enumerate(list_frequences):
            id_hex = str(hex(data["id"])).replace("0x","")
            id_hex = id_hex.upper()
            list_final.append({"id_dec": data["id"],
                                "id_hex": id_hex,
                                "t_moy": data["t_moy"],
                                "t_max": data["t_max"],
                                "t_min": data["t_min"],
                                })
def save_to_file():
    global list_final

    file_path = filedialog.asksaveasfilename(defaultextension=".xls")
    if file_path:
        # create a new Excel workbook
        workbook = openpyxl.Workbook()  
        # select the active worksheet
        worksheet = workbook.active

        # write the list to Excel
        worksheet.cell(row=1, column=1, value= "Id decimal")
        worksheet.cell(row=1, column=2, value= "Id hexa")
        worksheet.cell(row=1, column=3, value= "t_moy")
        worksheet.cell(row=1, column=4, value= "t_max")
        worksheet.cell(row=1, column=5, value= "t_min")
        for row_index, data in enumerate(list_final, start=2):
            worksheet.cell(row=row_index, column=1, value=data["id_dec"])
            worksheet.cell(row=row_index, column=2, value=data["id_hex"])
            worksheet.cell(row=row_index, column=3, value=data["t_moy"])
            worksheet.cell(row=row_index, column=4, value=data["t_max"])
            worksheet.cell(row=row_index, column=5, value=data["t_min"])
        workbook.save(file_path)

open_file()
get_time_data()
save_to_file()
print("File saved ✅")